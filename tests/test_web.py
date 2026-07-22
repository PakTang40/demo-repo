"""Smoke tests for the web layer.

Handlers are called directly rather than over a socket -- they are plain functions
of (conn, query, form), so this covers routing logic, form parsing, and every
template path without binding a port.
"""

import unittest

from apartment import db, money, repo
from apartment.excel import export_year
from apartment.web import finance, pages, server
from apartment.web.layout import CSS, esc


def form(**kwargs) -> dict[str, list[str]]:
    """Build a parsed-form dict the way parse_qs would."""
    return {key: [str(value)] for key, value in kwargs.items()}


class WebTestCase(unittest.TestCase):
    def setUp(self):
        self.conn = db.connect(":memory:")
        db.init_db(self.conn)
        db.seed_rooms(self.conn)

    def tearDown(self):
        self.conn.close()

    def occupy(self, code="101", name="สมชาย ใจดี"):
        result = server.post_move_in(
            self.conn, {}, form(code=code, full_name=name, phone="081-111-2222",
                                start_date="2026-07-01", monthly_rent="3500", deposit="")
        )
        self.assertIsInstance(result, server.Redirect)
        return result


class TestPagesRender(WebTestCase):
    def test_every_page_renders_for_an_empty_building(self):
        for render in (
            lambda: pages.dashboard_page(self.conn, "2026-07"),
            lambda: pages.rooms_page(self.conn),
            lambda: pages.meters_page(self.conn, "2026-07"),
            lambda: pages.invoices_page(self.conn, "2026-07"),
            lambda: finance.expenses_page(self.conn, "2026-07"),
            lambda: finance.reports_page(self.conn, 2026),
            lambda: finance.settings_page(self.conn),
        ):
            html = render()
            self.assertTrue(html.startswith("<!doctype html>"))
            self.assertIn("</html>", html)

    def test_room_grid_shows_all_thirty_rooms(self):
        html = pages.rooms_page(self.conn)
        for code in ("101", "110", "201", "310"):
            self.assertIn(f">{code}</div>", html)
        self.assertEqual(html.count('class="room '), 30)

    def test_room_grid_states_are_spelled_out_not_just_coloured(self):
        """Colour is decoration; the words are the actual status."""
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        repo.save_reading(self.conn, room["id"], "2026-07", 110, 2150,
                          water_prev=100, electric_prev=2000)
        repo.generate_invoices(self.conn, "2026-07", issue_date="2026-07-01")

        html = pages.rooms_page(self.conn)
        self.assertIn("ว่าง", html)          # the 29 empty rooms
        self.assertIn("รอออกบิล", html)      # ...or an unpaid one, depending on today
        self.assertIn('class="room-status"', html)
        # The legend explains each state with a count.
        self.assertIn('class="legend"', html)
        self.assertIn('class="sw vacant"', html)

    def test_room_grid_shows_the_outstanding_amount_on_the_plate(self):
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        repo.save_reading(self.conn, room["id"], "2026-06", 110, 2150,
                          water_prev=100, electric_prev=2000)
        repo.generate_invoices(self.conn, "2026-06", issue_date="2026-06-01")
        html = pages.rooms_page(self.conn)
        self.assertIn("ค้างชำระ", html)
        self.assertIn("4,880", html)
        self.assertIn('class="room owing"', html)

    def test_room_grid_summarises_the_building_at_the_top(self):
        self.occupy("101")
        html = pages.rooms_page(self.conn)
        self.assertIn("มีผู้เช่า", html)
        self.assertIn("1/30", html)
        self.assertIn("ห้องว่าง", html)
        self.assertIn("ค่าเช่ารวมต่อเดือน", html)

    def test_each_floor_gets_its_own_summary(self):
        html = pages.rooms_page(self.conn)
        self.assertEqual(html.count('class="floor-head"'), 3)
        for floor in (1, 2, 3):
            self.assertIn(f"ชั้น {floor}", html)

    def test_unknown_room_does_not_crash(self):
        html = pages.room_page(self.conn, "999")
        self.assertIn("ไม่พบห้อง", html)

    def test_vacant_and_occupied_plates_carry_the_two_agreed_fills(self):
        """The owner asked for this outright: empty = red plate, taken = green plate.

        It is the one visual rule on this page that is a business decision rather
        than a design preference, so it is pinned here. Note that red means "no
        income" on this board and "error" everywhere else -- see the CSS comment.
        """
        self.occupy("101")
        html = pages.rooms_page(self.conn)
        self.assertIn('class="room occupied"', html)   # 101 -- someone lives there
        self.assertIn('class="room vacant"', html)     # the other 29

        # Whole plate, not a stripe: the fill is set on .room itself, so the rule
        # that carries it is `.room.vacant`, not `.room.vacant .room-status`.
        vacant_rule = CSS.split(".room.vacant {")[1].split("}")[0]
        self.assertIn("--plan-vacant", vacant_rule)
        live_rule = CSS.split(".room.paid, .room.occupied, .room.owing")[1].split("}")[0]
        self.assertIn("--plan-live", live_rule)

    def test_arrears_stay_visible_on_a_green_plate(self):
        """A room that owes money is still occupied, so it keeps the green fill --
        which would hide the debt if amber did not override the status band."""
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        repo.save_reading(self.conn, room["id"], "2026-06", 110, 2150,
                          water_prev=100, electric_prev=2000)
        repo.generate_invoices(self.conn, "2026-06", issue_date="2026-06-01")
        html = pages.rooms_page(self.conn)

        self.assertIn('class="room owing"', html)
        owing_rule = CSS.split(".room.owing .room-status")[1].split("}")[0]
        self.assertIn("--plan-alert", owing_rule)

    def test_missing_invoice_does_not_crash(self):
        html = pages.invoice_page(self.conn, 4242)
        self.assertIn("ไม่พบใบแจ้งหนี้", html)

    def test_tenant_name_is_escaped(self):
        server.post_move_in(
            self.conn, {}, form(code="101", full_name='<script>alert("x")</script>',
                                start_date="2026-07-01", monthly_rent="3500", deposit="")
        )
        html = pages.rooms_page(self.conn)
        self.assertNotIn("<script>alert", html)
        self.assertIn("&lt;script&gt;", html)

    def test_esc_handles_none_and_quotes(self):
        self.assertEqual(esc(None), "")
        self.assertIn("&quot;", esc('a"b'))


class TestRouteTable(unittest.TestCase):
    def test_every_route_maps_to_a_callable(self):
        self.assertTrue(server.ROUTES)
        for (method, path), handler in server.ROUTES.items():
            self.assertIn(method, ("GET", "POST"), path)
            self.assertTrue(callable(handler), path)

    def test_nav_links_all_have_a_get_route(self):
        from apartment.web.layout import NAV

        for href, _label in NAV:
            self.assertIn(("GET", href), server.ROUTES, f"nav points at a missing route: {href}")


class TestFormHandlers(WebTestCase):
    def test_move_in_then_move_out(self):
        self.occupy("101")
        self.assertEqual(repo.room_by_code(self.conn, "101")["status"], "occupied")
        lease = repo.active_leases(self.conn)[0]
        server.post_move_out(
            self.conn, {}, form(lease_id=lease["id"], code="101",
                                ended_on="2026-12-31", deposit_refunded="7000")
        )
        self.assertEqual(repo.room_by_code(self.conn, "101")["status"], "vacant")

    def test_move_in_without_a_name_is_rejected(self):
        with self.assertRaises(server.AppError):
            server.post_move_in(self.conn, {}, form(code="101", full_name="", start_date="2026-07-01"))

    def test_blank_deposit_falls_back_to_the_default_months(self):
        self.occupy("101")
        lease = repo.active_leases(self.conn)[0]
        self.assertEqual(lease["deposit"], money.baht(3500) * 2)

    def test_meter_sheet_saves_only_filled_rows(self):
        self.occupy("101")
        self.occupy("102", name="สมหญิง")
        room101 = repo.room_by_code(self.conn, "101")
        room102 = repo.room_by_code(self.conn, "102")
        payload = {
            "period": ["2026-07"],
            "read_date": ["2026-07-01"],
            f"water_{room101['id']}": ["110"],
            f"electric_{room101['id']}": ["2150"],
            f"water_{room102['id']}": [""],      # left blank on purpose
            f"electric_{room102['id']}": [""],
        }
        server.post_meters(self.conn, {}, payload)
        readings = repo.readings_for(self.conn, "2026-07")
        self.assertIn(room101["id"], readings)
        self.assertNotIn(room102["id"], readings)

    def test_half_filled_meter_row_reports_an_error_instead_of_saving(self):
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        payload = {
            "period": ["2026-07"],
            f"water_{room['id']}": ["110"],
            f"electric_{room['id']}": [""],
        }
        result = server.post_meters(self.conn, {}, payload)
        self.assertIn("err=", result.location)
        self.assertEqual(repo.readings_for(self.conn, "2026-07"), {})

    def test_first_reading_without_an_opening_number_bills_zero_units(self):
        """The safe direction: under-bill once rather than bill the meter's whole life."""
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        server.post_meters(self.conn, {}, {
            "period": ["2026-07"], "read_date": ["2026-07-01"],
            f"water_{room['id']}": ["110"], f"electric_{room['id']}": ["2150"],
        })
        reading = repo.readings_for(self.conn, "2026-07")[room["id"]]
        self.assertEqual(reading["water_prev"], 110)
        self.assertEqual(reading["electric_prev"], 2150)

        server.post_generate_invoices(self.conn, {}, form(period="2026-07", issue_date="2026-07-01"))
        invoice = repo.invoices(self.conn, "2026-07")[0]
        self.assertEqual(invoice["total"], money.baht(3500))  # rent only

    def test_opening_reading_from_the_form_is_honoured(self):
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        server.post_meters(self.conn, {}, {
            "period": ["2026-07"], "read_date": ["2026-07-01"],
            f"water_prev_{room['id']}": ["100"], f"water_{room['id']}": ["110"],
            f"electric_prev_{room['id']}": ["2000"], f"electric_{room['id']}": ["2150"],
        })
        reading = repo.readings_for(self.conn, "2026-07")[room["id"]]
        self.assertEqual(reading["water_prev"], 100)
        self.assertEqual(reading["electric_prev"], 2000)

    def test_meter_sheet_offers_an_opening_field_only_on_a_first_reading(self):
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        html = pages.meters_page(self.conn, "2026-07")
        self.assertIn(f"water_prev_{room['id']}", html)
        self.assertIn("เดือนแรก", html)

        repo.save_reading(self.conn, room["id"], "2026-07", 110, 2150,
                          water_prev=100, electric_prev=2000)
        html = pages.meters_page(self.conn, "2026-08")
        self.assertNotIn(f"water_prev_{room['id']}", html)

    def test_full_monthly_cycle_through_handlers(self):
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        server.post_meters(self.conn, {}, {
            "period": ["2026-07"], "read_date": ["2026-07-01"],
            f"water_prev_{room['id']}": ["100"], f"water_{room['id']}": ["110"],
            f"electric_prev_{room['id']}": ["2000"], f"electric_{room['id']}": ["2150"],
        })
        server.post_generate_invoices(self.conn, {}, form(period="2026-07", issue_date="2026-07-01"))
        invoices = repo.invoices(self.conn, "2026-07")
        self.assertEqual(len(invoices), 1)
        # 3500 rent + 10 water x 18 + 150 electric x 8
        self.assertEqual(invoices[0]["total"], money.baht(4880))

        server.post_payment(self.conn, {}, form(
            invoice_id=invoices[0]["id"], amount="4880", paid_on="2026-07-05",
            method="transfer", period="2026-07"))
        detail = repo.invoice_detail(self.conn, invoices[0]["id"])
        self.assertEqual(detail["settlement"].status, "paid")

        # And the printable invoice renders with the payment on it.
        html = pages.invoice_page(self.conn, invoices[0]["id"])
        self.assertIn("INV-2026-07-101", html)
        self.assertIn("ชำระแล้ว", html)

    def test_generate_with_no_readings_redirects_with_an_error(self):
        self.occupy("101")
        result = server.post_generate_invoices(self.conn, {}, form(period="2026-07"))
        self.assertIn("err=", result.location)

    def test_expense_requires_an_amount(self):
        with self.assertRaises(server.AppError):
            server.post_expense(self.conn, {}, form(
                spent_on="2026-07-10", category="repair", description="ซ่อม", amount=""))

    def test_expense_is_recorded_and_shows_on_the_page(self):
        server.post_expense(self.conn, {}, form(
            spent_on="2026-07-10", category="repair", description="ซ่อมปั๊มน้ำ",
            amount="3000", room_id="", vendor="ช่างสมศักดิ์", period="2026-07"))
        html = finance.expenses_page(self.conn, "2026-07")
        self.assertIn("ซ่อมปั๊มน้ำ", html)
        self.assertIn("3,000.00", html)

    def test_settings_round_trip_baht_to_satang(self):
        server.post_settings(self.conn, {}, form(
            building_name="หอพักบ้านสวน", water_rate="20", electric_rate="7.50",
            due_day="10", common_fee="0", base_rent="3500", deposit_months="2",
            late_fee_per_day="50", late_fee_grace_days="3",
            building_address="", promptpay_id="0812345678"))
        settings = db.get_settings(self.conn)
        self.assertEqual(settings["building_name"], "หอพักบ้านสวน")
        self.assertEqual(int(settings["water_rate"]), money.baht(20))
        self.assertEqual(int(settings["electric_rate"]), money.baht("7.50"))
        # The new name reaches the page header.
        self.assertIn("หอพักบ้านสวน", pages.rooms_page(self.conn))

    def test_bad_money_input_raises_a_readable_error(self):
        with self.assertRaises(server.AppError):
            server.post_expense(self.conn, {}, form(
                spent_on="2026-07-10", category="repair", description="x", amount="abc"))


class TestExcelExport(WebTestCase):
    def test_workbook_has_the_expected_sheets(self):
        self.occupy("101")
        room = repo.room_by_code(self.conn, "101")
        repo.save_reading(self.conn, room["id"], "2026-07", 110, 2150,
                          water_prev=100, electric_prev=2000)
        repo.generate_invoices(self.conn, "2026-07", issue_date="2026-07-01")
        invoice = repo.invoices(self.conn, "2026-07")[0]
        repo.record_payment(self.conn, invoice["id"], money.baht(4880), paid_on="2026-07-05")
        repo.record_expense(self.conn, "2026-07-10", "repair", "ซ่อมปั๊มน้ำ", money.baht(3000))

        payload = export_year(self.conn, 2026)
        self.assertTrue(payload.startswith(b"PK"))  # a real xlsx is a zip

        import io

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(payload))
        self.assertEqual(
            workbook.sheetnames,
            ["สรุปรายเดือน", "ใบแจ้งหนี้", "การชำระเงิน", "รายจ่าย", "ผลตอบแทนรายห้อง"],
        )
        # July collected 4,880 baht -- written as a number, not a string.
        july = workbook["สรุปรายเดือน"]["C8"].value
        self.assertAlmostEqual(july, 4880.0)
        self.assertIsInstance(july, (int, float))

    def test_export_of_an_empty_year_still_produces_a_workbook(self):
        payload = export_year(self.conn, 2030)
        self.assertTrue(payload.startswith(b"PK"))


if __name__ == "__main__":
    unittest.main()
