"""Excel export.

The owner already works in Excel, so a year's records must be able to leave this
system whole -- as numbers, not as a picture of numbers. Every money cell is written
as a real float in baht with a Thai currency format, so it sums correctly in Excel.
"""

from __future__ import annotations

import io
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import money, reports
from .web.finance import CATEGORY_LABEL, MONTH_TH

BAHT_FORMAT = "#,##0.00"
HEADER_FILL = PatternFill("solid", fgColor="1F5F4E")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"


def _autosize(sheet, minimum: int = 10, maximum: int = 42) -> None:
    for column in sheet.columns:
        width = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(
            max(width + 3, minimum), maximum
        )


def _money_column(sheet, *columns: str) -> None:
    for letter in columns:
        for cell in sheet[letter][1:]:
            cell.number_format = BAHT_FORMAT


def export_year(conn: sqlite3.Connection, year: int) -> bytes:
    """A five-sheet workbook covering one calendar year."""
    workbook = Workbook()

    # -- สรุปรายเดือน -----------------------------------------------------
    summary_sheet = workbook.active
    summary_sheet.title = "สรุปรายเดือน"
    _write_header(
        summary_sheet,
        ["เดือน", "ออกบิล", "เก็บได้", "รายจ่าย", "กำไรสุทธิ", "อัตราเก็บเงิน",
         "ห้องมีผู้เช่า", "ห้องทั้งหมด", "อัตราการเช่า"],
    )
    table = reports.yearly_table(conn, year)
    for index, month in enumerate(table):
        summary_sheet.append([
            MONTH_TH[index],
            float(money.to_baht(month.invoiced)),
            float(money.to_baht(month.collected)),
            float(money.to_baht(month.expenses)),
            float(money.to_baht(month.net)),
            round(month.collection_rate, 4),
            month.rooms_occupied,
            month.rooms_total,
            round(month.occupancy_rate, 4),
        ])
    last = summary_sheet.max_row
    summary_sheet.append([
        "รวมทั้งปี",
        f"=SUM(B2:B{last})", f"=SUM(C2:C{last})", f"=SUM(D2:D{last})", f"=SUM(E2:E{last})",
        None, None, None, None,
    ])
    for cell in summary_sheet[summary_sheet.max_row]:
        cell.font = Font(bold=True)
    _money_column(summary_sheet, "B", "C", "D", "E")
    for letter in ("F", "I"):
        for cell in summary_sheet[letter][1:]:
            cell.number_format = "0.0%"
    _autosize(summary_sheet)

    # -- ใบแจ้งหนี้ --------------------------------------------------------
    invoice_sheet = workbook.create_sheet("ใบแจ้งหนี้")
    _write_header(
        invoice_sheet,
        ["งวด", "เลขที่", "ห้อง", "ชั้น", "ผู้เช่า", "วันที่ออก", "ครบกำหนด",
         "ค่าเช่า", "ค่าน้ำ", "ค่าไฟ", "ค่าส่วนกลาง", "ค่าปรับ", "อื่น ๆ",
         "ยอดรวม", "ชำระแล้ว", "คงค้าง"],
    )
    for row in conn.execute(
        """
        SELECT i.period, i.number, r.code AS room_code, r.floor, t.full_name, i.issue_date,
               i.due_date, i.total,
               COALESCE((SELECT SUM(p.amount) FROM payment p WHERE p.invoice_id = i.id), 0) AS paid,
               COALESCE((SELECT SUM(amount) FROM invoice_line WHERE invoice_id = i.id AND kind='rent'), 0) AS rent,
               COALESCE((SELECT SUM(amount) FROM invoice_line WHERE invoice_id = i.id AND kind='water'), 0) AS water,
               COALESCE((SELECT SUM(amount) FROM invoice_line WHERE invoice_id = i.id AND kind='electric'), 0) AS electric,
               COALESCE((SELECT SUM(amount) FROM invoice_line WHERE invoice_id = i.id AND kind='common'), 0) AS common,
               COALESCE((SELECT SUM(amount) FROM invoice_line WHERE invoice_id = i.id AND kind='late_fee'), 0) AS late_fee,
               COALESCE((SELECT SUM(amount) FROM invoice_line WHERE invoice_id = i.id AND kind='other'), 0) AS other
        FROM invoice i
        JOIN room r   ON r.id = i.room_id
        JOIN lease l  ON l.id = i.lease_id
        JOIN tenant t ON t.id = l.tenant_id
        WHERE i.status = 'issued' AND substr(i.period, 1, 4) = ?
        ORDER BY i.period, r.floor, r.code
        """,
        (str(year),),
    ):
        invoice_sheet.append([
            row["period"], row["number"], row["room_code"], row["floor"], row["full_name"],
            row["issue_date"], row["due_date"],
            float(money.to_baht(row["rent"])), float(money.to_baht(row["water"])),
            float(money.to_baht(row["electric"])), float(money.to_baht(row["common"])),
            float(money.to_baht(row["late_fee"])), float(money.to_baht(row["other"])),
            float(money.to_baht(row["total"])), float(money.to_baht(row["paid"])),
            float(money.to_baht(row["total"] - row["paid"])),
        ])
    _money_column(invoice_sheet, "H", "I", "J", "K", "L", "M", "N", "O", "P")
    _autosize(invoice_sheet)

    # -- การชำระเงิน -------------------------------------------------------
    payment_sheet = workbook.create_sheet("การชำระเงิน")
    _write_header(payment_sheet, ["วันที่ชำระ", "ห้อง", "ผู้เช่า", "งวด", "เลขที่บิล",
                                  "ช่องทาง", "อ้างอิง", "จำนวนเงิน"])
    for row in conn.execute(
        """
        SELECT p.paid_on, r.code AS room_code, t.full_name, i.period, i.number,
               p.method, p.reference, p.amount
        FROM payment p
        JOIN invoice i ON i.id = p.invoice_id
        JOIN room r    ON r.id = i.room_id
        JOIN lease l   ON l.id = i.lease_id
        JOIN tenant t  ON t.id = l.tenant_id
        WHERE substr(p.paid_on, 1, 4) = ?
        ORDER BY p.paid_on, r.code
        """,
        (str(year),),
    ):
        payment_sheet.append([
            row["paid_on"], row["room_code"], row["full_name"], row["period"], row["number"],
            row["method"], row["reference"], float(money.to_baht(row["amount"])),
        ])
    _money_column(payment_sheet, "H")
    _autosize(payment_sheet)

    # -- รายจ่าย -----------------------------------------------------------
    expense_sheet = workbook.create_sheet("รายจ่าย")
    _write_header(expense_sheet, ["วันที่", "หมวด", "รายละเอียด", "ห้อง", "ผู้ขาย", "จำนวนเงิน"])
    for row in conn.execute(
        """
        SELECT e.spent_on, e.category, e.description, r.code AS room_code, e.vendor, e.amount
        FROM expense e
        LEFT JOIN room r ON r.id = e.room_id
        WHERE substr(e.spent_on, 1, 4) = ?
        ORDER BY e.spent_on
        """,
        (str(year),),
    ):
        expense_sheet.append([
            row["spent_on"], CATEGORY_LABEL.get(row["category"], row["category"]),
            row["description"], row["room_code"] or "ทั้งตึก", row["vendor"],
            float(money.to_baht(row["amount"])),
        ])
    _money_column(expense_sheet, "F")
    _autosize(expense_sheet)

    # -- ผลตอบแทนรายห้อง ---------------------------------------------------
    room_sheet = workbook.create_sheet("ผลตอบแทนรายห้อง")
    _write_header(room_sheet, ["ห้อง", "ชั้น", "สถานะ", "เดือนที่ออกบิล", "ออกบิล",
                               "เก็บได้", "คงค้าง"])
    for row in reports.room_performance(conn, year):
        room_sheet.append([
            row["room_code"], row["floor"], row["status"], row["months_billed"],
            float(money.to_baht(row["invoiced"])), float(money.to_baht(row["collected"])),
            float(money.to_baht(row["outstanding"])),
        ])
    _money_column(room_sheet, "E", "F", "G")
    _autosize(room_sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
